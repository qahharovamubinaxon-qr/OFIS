"""АМИНА — a worker's account in the office's own app for migrants.

The office already has a working importer at ``C:\\src\\amina_admin_import``:
a vertical Excel of twenty-one ``field``/``value`` rows, and ``npm run
import``, which creates the Firebase account and hangs the worker's documents
off it. What was done by hand — cropping five photographs, uploading them
somewhere public, pasting the links into the right rows, retyping the passport
— is what this does instead.

**Their folder is not ours.** ``import_users.js``, ``package.json``,
``serviceAccountKey.json`` and ``node_modules`` are the office's working
machinery and are never written to, moved or read for their contents. The one
file touched is the Excel, and only column B of its ``user`` sheet — every
row label, both sheets, and the text formatting stay exactly as they were.
The untouched original is copied into OFIS's own AppData the first time, so
there is always a way back that does not depend on us being right.

Two rules the office set, kept literally
----------------------------------------
The password IS the worker's phone with ``+7`` written as ``8`` — so
``+7 996 681-84-92`` becomes ``89966818492``. The login is his surname in
latin letters with the last four digits of that phone: ``sharipov8492``.

What the passport gives and what is typed
-----------------------------------------
Read: the name, the date of birth, the country of birth, the citizenship.
Typed by the office: the phone, the address, and the card's number and expiry.
Everything lands in an editable box first — nothing is sent anywhere until the
operator has looked at it.
"""

from __future__ import annotations

import io
import json
import os
import re
import shutil
import subprocess
from dataclasses import dataclass, field
from pathlib import Path

from src.common.errors import OfisError, ValidationError
from src.common.logging import get_logger
from src.config import paths

log = get_logger(__name__)

SECTION = "amina"

#: The office's own importer. Ours to read from and to put an Excel into —
#: never to modify otherwise.
DEFAULT_FOLDER = Path(r"C:\src\amina_admin_import")
EXCEL_NAME = "amina_users_template_vertical.xlsx"
SHEET = "user"
#: Where the operator's typed folder is kept, if it is ever not the default.
KEY_FOLDER = "amina.folder"

# --------------------------------------------------------------- the rows
#: Every row of the Excel, in the order the office's own file has them. The
#: importer reads by label, not by position, but a file that still looks like
#: the one they know is a file they can open and check.
FIELDS = (
    "email", "password", "fullName", "gender", "dob", "birthPlace",
    "citizenship", "phone", "extraPhone", "addressStreet", "addressHouse",
    "addressApartment", "addressSettlement", "kigNumber", "kigExpire",
    "identityUrl", "innUrl", "dmsUrl", "educationUrl", "patentUrl",
)

#: The five document rows, and what the app calls each one. Taken from the
#: office's own ``import_users.js`` so the two cannot drift apart.
DOCS = ("identityUrl", "innUrl", "dmsUrl", "educationUrl", "patentUrl")
DOC_NAMES = {
    "identityUrl": "Паспорт",
    "innUrl": "ИНН",
    "dmsUrl": "DMS",
    "educationUrl": "Сертификат",
    "patentUrl": "Патент",
}
#: Identity takes two pictures — an ID-card passport has a front and a back,
#: and the office wants both on ONE sheet. The rest take one each.
DOC_LIMIT = {"identityUrl": 2}


def doc_limit(key: str) -> int:
    return DOC_LIMIT.get(key, 1)


# ------------------------------------------------------------ the folder
def folder(settings=None) -> Path:
    """Where the office's importer lives."""
    if settings is not None:
        typed = str(settings.get(KEY_FOLDER, "") or "").strip()
        if typed:
            return Path(typed)
    return DEFAULT_FOLDER


def excel_path(settings=None) -> Path:
    return folder(settings) / EXCEL_NAME


def _our_folder() -> Path:
    made = paths.user_templates_dir() / SECTION
    made.mkdir(parents=True, exist_ok=True)
    return made


def check_folder(settings=None) -> None:
    """Say what is missing BEFORE a passport is read, not after."""
    where = folder(settings)
    if not where.exists():
        raise ValidationError(
            f"АМИНА папкаси топилмади: {where}\n"
            "Созламаларда папка йўлини кўрсатинг.")
    if not excel_path(settings).exists():
        raise ValidationError(f"Эксел топилмади: {excel_path(settings)}")
    if not (where / "import_users.js").exists():
        raise ValidationError(f"import_users.js топилмади: {where}")
    if shutil.which("npm") is None:
        raise ValidationError(
            "npm топилмади — Node.js ўрнатилганини текширинг.")


def excel_is_open(settings=None) -> bool:
    """Is the Excel actually held open — can we write to it or not?

    NOT by looking for the ``~$`` file Excel leaves beside an open workbook.
    That file survives a crash, and the office has a stale one sitting in
    their folder right now, months older than the workbook itself. Refusing
    on its account would refuse for ever.

    So the real question is asked instead: open the file for writing. Windows
    denies that while Excel holds it, and grants it the moment Excel lets go
    — which is exactly the thing worth knowing, because a workbook written
    under a live Excel is thrown away the instant the office saves.
    """
    path = excel_path(settings)
    if not path.exists():
        return False
    try:
        with path.open("r+b"):
            return False
    except PermissionError:
        return True
    except OSError:
        return False


# ------------------------------------------------- the login and password
_LATIN = {
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "yo",
    "ж": "j", "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m",
    "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
    "ф": "f", "х": "kh", "ц": "ts", "ч": "ch", "ш": "sh", "щ": "sch",
    "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
    "ғ": "g", "қ": "q", "ҳ": "h", "ў": "o",
}


def latin(word: str) -> str:
    """«Шарипов» → «sharipov». Letters only — a login has nothing else in it."""
    out = "".join(_LATIN.get(ch, ch) for ch in (word or "").strip().lower())
    return "".join(ch for ch in out if ch.isascii() and ch.isalpha())


def digits(phone: str) -> str:
    return re.sub(r"\D", "", phone or "")


def password_of(phone: str) -> str:
    """«+7 996 681-84-92» → «89966818492». The office's rule, kept literally.

    A Russian number is eleven digits written with either a 7 or an 8 in
    front, and the office writes it with an 8. A number that is not eleven
    digits is handed back as its digits alone: inventing a shape for it would
    hand a worker a password that does not open his account.
    """
    only = digits(phone)
    if len(only) == 11 and only[0] == "7":
        return "8" + only[1:]
    if len(only) == 10:
        return "8" + only
    return only


def email_of(surname: str, phone: str, domain: str = "gmail.com") -> str:
    """«Шарипов» + «…84-92» → «sharipov8492@gmail.com»."""
    stem = latin(surname)
    tail = digits(phone)[-4:]
    if not stem:
        return ""
    return f"{stem}{tail}@{domain}"


# ------------------------------------------------------- what the passport says
#: How the app writes a citizenship, against how a passport prints a country.
#: Only the ones the office actually deals with; anything else falls through
#: to «Республика X», which the operator can correct on screen.
_CITIZENSHIP = {
    "таджикистан": "Республика Таджикистан",
    "узбекистан": "Республика Узбекистан",
    "киргизия": "Кыргызская Республика",
    "кыргызстан": "Кыргызская Республика",
    "казахстан": "Республика Казахстан",
    "туркменистан": "Туркменистан",
    "азербайджан": "Азербайджанская Республика",
    "армения": "Республика Армения",
    "молдова": "Республика Молдова",
    "беларусь": "Республика Беларусь",
    "украина": "Украина",
    "россия": "Российская Федерация",
}


def citizenship_of(country: str) -> str:
    said = (country or "").strip()
    if not said:
        return ""
    return _CITIZENSHIP.get(said.lower(), f"Республика {said.title()}")


@dataclass
class AminaData:
    """One worker, exactly as the twenty-one rows want him."""

    # read from the passport
    full_name: str = ""
    gender: str = ""                      # «Мужской» / «Женский»
    dob: str = ""                         # DD-MM-YYYY, as the office's file
    birth_place: str = ""
    citizenship: str = ""
    # typed by the office
    phone: str = ""
    extra_phone: str = ""
    street: str = ""
    house: str = ""
    apartment: str = ""
    settlement: str = ""
    city: str = "Москва"
    kig_number: str = ""
    kig_expire: str = ""                  # YYYY-MM-DD, as the office's file
    #: field name → the direct links already uploaded, if any
    urls: dict[str, str] = field(default_factory=dict)

    def login(self) -> str:
        return email_of(self.full_name.split()[0] if self.full_name else "",
                        self.phone)

    def password(self) -> str:
        return password_of(self.phone)

    def address_line(self) -> str:
        """«Москва, улица Беловежская, 71» — the office's own shape."""
        if self.settlement.strip():
            return self.settlement.strip()
        parts = [p.strip() for p in (self.city, self.street, self.house)
                 if p and p.strip()]
        return ", ".join(parts)

    def rows(self) -> dict[str, str]:
        """The twenty-one values, ready for column B."""
        made = {
            "email": self.login(),
            "password": self.password(),
            "fullName": self.full_name.strip(),
            "gender": self.gender.strip(),
            "dob": self.dob.strip(),
            "birthPlace": self.birth_place.strip(),
            "citizenship": self.citizenship.strip(),
            "phone": self.phone.strip(),
            "extraPhone": (self.extra_phone or self.phone).strip(),
            "addressStreet": self.street.strip(),
            "addressHouse": self.house.strip(),
            "addressApartment": self.apartment.strip(),
            "addressSettlement": self.address_line(),
            "kigNumber": self.kig_number.strip(),
            "kigExpire": self.kig_expire.strip(),
        }
        for key in DOCS:
            made[key] = (self.urls.get(key) or "").strip()
        return made


def data_of(passport) -> AminaData:
    """What the passport gives. The rest is typed, and all of it editable."""
    said = getattr(getattr(passport, "gender", None), "value",
                   getattr(passport, "gender", "")) or ""
    born = getattr(passport, "birth_date", None)
    country = (getattr(passport, "nationality", "") or "").strip()
    where = (getattr(passport, "birth_place", "") or "").strip()
    names = [(getattr(passport, part, "") or "").title()
             for part in ("surname", "name", "patronymic")]
    return AminaData(
        full_name=" ".join(p for p in names if p),
        gender="Женский" if str(said).lower().startswith(("f", "ж")) else "Мужской",
        dob=born.strftime("%d-%m-%Y") if born else "",
        birth_place=(where or country).title(),
        citizenship=citizenship_of(country))


# ------------------------------------------------------------- the sheet
#: How large the finished picture is made. The app does not enlarge one
#: smaller than its frame, and the office's own uploads are 2480×3507, so
#: anything under this sits in the middle of the frame looking shrunken.
LONG_SIDE = 3500
#: A white hairline between two documents sharing a sheet — enough to see
#: where one ends, not enough to read as empty space.
GUTTER = 0.015

#: How much more textured than its background a document is. Measured off the
#: office's own passport photograph: the wall behind it scored 3.5, the
#: passport 19.0. Nine sits well clear of the paper and well under the print.
_TEXTURE = 9.0
#: A crop smaller than this share of the picture is not the document — it is
#: a stamp, a photograph, a logo printed on it. Those are left alone.
_MIN_SHARE = 0.25


def _tighten(rgb):
    """The document alone: the background found by TEXTURE, and cut away.

    Brightness cannot do this. In the office's own photograph the desk is
    BRIGHTER than the passport (168 against 148), and its colour is no help
    either — a beige wall is slightly more saturated than a grey-pink
    passport page. What separates them is detail: printing, guilloche and
    the photograph give the document local variation everywhere, and a bare
    surface has almost none.

    So: local standard deviation, threshold, the largest region that leaves,
    its minimum-area rectangle — which also gives the angle it was lying at,
    so the crop straightens it at the same time.

    A region too small to be the document is refused and the picture is
    handed back whole. Cutting a passport down to its own portrait photo
    would be far worse than leaving a margin of desk around it.
    """
    import cv2
    import numpy as np

    height, width = rgb.shape[:2]
    grey = cv2.cvtColor(rgb, cv2.COLOR_RGB2GRAY).astype(np.float32)
    mean = cv2.blur(grey, (15, 15))
    texture = np.sqrt(np.maximum(cv2.blur(grey ** 2, (15, 15)) - mean ** 2, 0))

    mask = (texture > _TEXTURE).astype(np.uint8) * 255
    close = max(9, int(min(height, width) * 0.03)) | 1
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, np.ones((close, close), np.uint8))
    opening = max(5, close // 2) | 1
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, np.ones((opening, opening), np.uint8))

    found, labels, stats, _ = cv2.connectedComponentsWithStats(mask, 8)
    if found < 2:
        return rgb
    biggest = 1 + int(np.argmax(stats[1:, cv2.CC_STAT_AREA]))
    if stats[biggest, cv2.CC_STAT_AREA] < _MIN_SHARE * height * width:
        log.info("АМИНА: ҳужжат чегараси ишончсиз — расм бутун қолди")
        return rgb

    (mid_x, mid_y), (box_w, box_h), angle = cv2.minAreaRect(
        cv2.findNonZero((labels == biggest).astype(np.uint8)))
    if angle < -45:
        angle, box_w, box_h = angle + 90, box_h, box_w
    if abs(angle) > 15:                       # not a tilt — a bad reading
        angle = 0.0
    turned = cv2.warpAffine(
        rgb, cv2.getRotationMatrix2D((mid_x, mid_y), angle, 1.0),
        (width, height), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
    left, top = max(0, int(mid_x - box_w / 2)), max(0, int(mid_y - box_h / 2))
    right = min(width, int(mid_x + box_w / 2))
    bottom = min(height, int(mid_y + box_h / 2))
    if right - left < 32 or bottom - top < 32:
        return rgb
    return turned[top:bottom, left:right]


def cut(image: bytes):
    """One photograph → the document alone: found, straightened, finished.

    Three steps, each somebody else's job. `scan_one` finds the document and
    squares it up; `_tighten` takes off whatever desk it missed; and
    `doc_enhance.finish` does what the office's own scanner at
    qrixtools.com does — evens out the shadow the phone cast and sets the
    picture to the document's real shape at 300 dpi.
    """
    import numpy as np

    from src.services.doc_enhance import finish
    from src.services.doc_scan_service import scan_one

    page = _tighten(scan_one(image, grayscale=False))
    return np.asarray(finish(page, min_long=LONG_SIDE).convert("RGB"))


def sheet_jpeg(images: list[bytes]) -> bytes:
    """One or two documents, cut out, as ONE picture with nothing else in it.

    There is no white page under this any more. The office was plain about
    it — «полный саҳифа катакни эгалайдиган қилибер, бўш жой қолмасин» —
    and a page is exactly what was making the empty space: a passport spread
    is taller than A4, so fitting it onto one left a white band down either
    side, and the desk it was photographed on filled the rest.

    So the picture IS the document. Every pixel of it is passport, and the
    app has nothing to letterbox.

    Two pictures still share one sheet, one above the other with a hairline
    between them — an ID-card passport's front and back are one document.
    They are matched to the same width so neither looks the larger.
    """
    from PIL import Image

    if not images:
        raise ValidationError("Камида битта расм керак.")
    if len(images) > 2:
        raise ValidationError("Битта майдонга кўпи билан 2 та расм.")

    cuts = [Image.fromarray(cut(one)) for one in images]

    if len(cuts) == 2:
        width = max(one.width for one in cuts)
        scaled = [one.resize((width, max(1, round(one.height * width / one.width))),
                             Image.LANCZOS) for one in cuts]
        gutter = max(2, int(width * GUTTER))
        sheet = Image.new("RGB", (width, sum(s.height for s in scaled) + gutter),
                          "white")
        sheet.paste(scaled[0], (0, 0))
        sheet.paste(scaled[1], (0, scaled[0].height + gutter))
    else:
        sheet = cuts[0]

    scale = LONG_SIDE / max(sheet.width, sheet.height)
    if scale > 1:                     # never shrink a picture that is already big
        sheet = sheet.resize((max(1, round(sheet.width * scale)),
                              max(1, round(sheet.height * scale))), Image.LANCZOS)

    buf = io.BytesIO()
    sheet.save(buf, format="JPEG", quality=90, optimize=True)
    return buf.getvalue()


# ------------------------------------------------------------- the Excel
def _backup(source: Path) -> None:
    """Keep the office's own file, once, somewhere that is ours.

    Not beside theirs: their folder is not a place we leave things in.
    """
    kept = _our_folder() / f"original_{EXCEL_NAME}"
    if not kept.exists() and source.exists():
        shutil.copyfile(source, kept)
        log.info("АМИНА: асл эксел сақланди — %s", kept)


def fill_excel(values: dict[str, str], path: Path) -> list[str]:
    """Write column B and nothing else. Returns the rows actually written.

    Loaded and saved with openpyxl, so the row labels, the second sheet and
    the text formatting the office set all survive: only the value cell of a
    row we have something for is assigned, and every cell stays a STRING.
    That last part matters — a date written as a date reaches the importer as
    a serial number, and the worker's birthday would land in the app as
    «44911».
    """
    from openpyxl import load_workbook

    if not path.exists():
        raise ValidationError(f"Эксел топилмади: {path}")
    _backup(path)

    book = load_workbook(path)
    if SHEET not in book.sheetnames:
        raise ValidationError(f"Эксел ичида «{SHEET}» варағи йўқ")
    sheet = book[SHEET]

    written: list[str] = []
    for row in sheet.iter_rows(min_row=2):
        label = row[0].value
        if not label or str(label).strip() not in values:
            continue
        name = str(label).strip()
        cell = row[1]
        cell.value = str(values[name] or "")
        cell.number_format = "@"
        written.append(name)

    missing = [f for f in values if f not in written]
    if missing:
        raise ValidationError(
            "Эксел бу қаторларни ўзида сақламайди: " + ", ".join(missing))

    book.save(path)
    log.info("АМИНА: эксел тўлдирилди — %d қатор", len(written))
    return written


def read_excel(path: Path) -> dict[str, str]:
    """What the Excel says now — for showing the office what was written."""
    from openpyxl import load_workbook

    book = load_workbook(path, data_only=True)
    sheet = book[SHEET]
    out: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            out[str(row[0]).strip()] = "" if row[1] is None else str(row[1])
    return out


# ---------------------------------------------------------- the importer
#: What the office's own script prints when it has finished.
_DONE = "IMPORT FINISHED"


def run_import(where: Path, timeout: int = 600) -> str:
    """``npm install`` then ``npm run import``, in the office's own folder.

    Both, in that order, because that is what their note says. Their files are
    read and executed, never written to. The whole output comes back so a
    failure — an e-mail already registered, a key expired — is shown to the
    office rather than swallowed.
    """
    npm = shutil.which("npm")
    if npm is None:
        raise OfisError("npm топилмади — Node.js ўрнатилганини текширинг.")

    quiet = {**os.environ, "npm_config_loglevel": "error", "NO_COLOR": "1"}
    said: list[str] = []
    for step in (["install"], ["run", "import"]):
        shown = "npm " + " ".join(step)
        said.append(f"$ {shown}")
        try:
            done = subprocess.run(                       # noqa: S603
                [npm, *step], cwd=str(where), env=quiet, timeout=timeout,
                capture_output=True, text=True, encoding="utf-8",
                errors="replace", shell=False)
        except subprocess.TimeoutExpired as exc:
            raise OfisError(f"«{shown}» {timeout} сонияда тугамади") from exc
        said.append((done.stdout or "").strip())
        if (done.stderr or "").strip():
            said.append((done.stderr or "").strip())
        if done.returncode != 0:
            raise OfisError(f"«{shown}» хато билан тугади:\n"
                            + "\n".join(s for s in said if s))

    out = "\n".join(s for s in said if s)
    if _DONE not in out:
        raise OfisError("Импорт тугамади — Firebase жавоби:\n" + out)
    return out


# -------------------------------------------------------------- the making
@dataclass(frozen=True)
class AminaResult:
    """What the office hands the worker, and what the run said."""

    login: str
    password: str
    #: field name → the direct link that went into the Excel
    urls: dict[str, str]
    output: str

    def slip(self) -> str:
        """The two lines the worker needs, and nothing else."""
        return f"Логин: {self.login}\nПарол: {self.password}"


class AminaService:
    """Pictures up, Excel filled, importer run — in that order, once."""

    def __init__(self, settings=None) -> None:
        self._settings = settings

    @property
    def settings(self):
        """Where the folder and the imgbb key are read from."""
        return self._settings

    def _key(self) -> str:
        from src.services.imgbb import KEY_IMGBB

        if self._settings is None:
            return ""
        return str(self._settings.get(KEY_IMGBB, "") or "").strip()

    def upload(self, key: str, images: list[bytes]) -> str:
        """One field's pictures → one A4 sheet → its direct link."""
        from src.services.imgbb import upload

        link = upload(sheet_jpeg(images), self._key(), name=key)
        log.info("АМИНА: %s — %s", key, link)
        return link

    def create(self, data: AminaData,
               images: dict[str, list[bytes]] | None = None,
               *, run: bool = True) -> AminaResult:
        """The whole run. Nothing leaves this machine until it all checks out.

        The order is deliberate: everything that can be refused is refused
        first, then the pictures go up, then the Excel, then the importer.
        A worker half-created — an account with no documents, or documents
        with no account — is worse than one not created at all.
        """
        where = folder(self._settings)
        check_folder(self._settings)
        if excel_is_open(self._settings):
            raise ValidationError(
                "Эксел ҳозир Excel'да очиқ турибди — ёпинг ва қайта уриниб "
                "кўринг (акс ҳолда ёзилгани сақланмайди).")
        if not data.full_name.strip():
            raise ValidationError("ФИО керак — паспортни ўқитинг")
        if not digits(data.phone):
            raise ValidationError("Телефон рақами керак — парол ундан ясалади")
        if not data.login():
            raise ValidationError("Логин ясалмади — фамилияни текширинг")

        sending = {k: v for k, v in (images or {}).items() if v}
        if sending and not self._key():
            raise ValidationError(
                "imgbb API калити йўқ — Sozlamalar'да «КРКОД РЕГ» қисмига "
                "калитни киритинг.")

        urls = dict(data.urls)
        for key in DOCS:                     # in the Excel's own order
            if key in sending:
                urls[key] = self.upload(key, sending[key])
        data.urls = urls

        fill_excel(data.rows(), excel_path(self._settings))
        output = run_import(where) if run else ""
        log.info("АМИНА: %s — аккаунт очилди", data.login())
        return AminaResult(login=data.login(), password=data.password(),
                           urls={k: v for k, v in urls.items() if v},
                           output=output)


# ------------------------------------------------------- what was typed
_TYPED = "typed.json"
#: The office types the same city and the same second phone all day.
REMEMBERED = ("city", "extra_phone")


def typed() -> dict[str, str]:
    store = _our_folder() / _TYPED
    if not store.exists():
        return {}
    try:
        kept = json.loads(store.read_text("utf-8"))
    except (OSError, ValueError):
        return {}
    return {k: str(kept.get(k, "") or "") for k in REMEMBERED} \
        if isinstance(kept, dict) else {}


def remember_typed(**boxes: str) -> None:
    kept = typed()
    for key, value in boxes.items():
        if key in REMEMBERED:
            kept[key] = (value or "").strip()
    (_our_folder() / _TYPED).write_text(
        json.dumps(kept, ensure_ascii=False), "utf-8")


__all__ = ["DOCS", "DOC_LIMIT", "DOC_NAMES", "FIELDS", "AminaData",
           "AminaResult", "AminaService", "check_folder", "citizenship_of",
           "data_of", "digits", "doc_limit", "email_of", "excel_is_open",
           "excel_path", "fill_excel", "folder", "latin", "password_of",
           "read_excel", "remember_typed", "run_import", "sheet_jpeg",
           "typed"]
