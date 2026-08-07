"""АМИНА — the worker's account in the office's own app.

What is checked here is what the office was exact about: the password is the
phone with an 8 in front of it, the login is the surname and the last four
digits, however many pictures are sent is however many rows get a link, an
ID-card's two sides share ONE sheet — and, above all, that the office's own
folder comes out of a run with nothing but its Excel touched.

The Excel is checked against the office's real file, filled by hand, row by
row: the program has to reproduce it exactly, or the importer downstream will
read something the office never meant.
"""

from __future__ import annotations

import hashlib
import io
import shutil
import tempfile
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import pytest
from src.common.errors import OfisError, ValidationError
from src.config import paths
from src.services import amina_service as store

#: The office's own filled sample. Every value here was typed by them.
SAMPLE = {
    "email": "sharipov8492@gmail.com",
    "password": "89966818492",
    "fullName": "Шарипов Эхромиддин Талбишоевич",
    "gender": "Мужской",
    "dob": "20-01-1998",
    "birthPlace": "Таджикистан",
    "citizenship": "Республика Таджикистан",
    "phone": "+7 996 681-84-92",
    "extraPhone": "+7 996 681-84-92",
    "addressStreet": "улица Беловежская",
    "addressHouse": "71",
    "addressApartment": "94",
    "addressSettlement": "Москва, улица Беловежская, 71",
    "kigNumber": "АВ0461171",
    "kigExpire": "2031-06-05",
    "identityUrl": "https://i.ibb.co/FqWfdgRT/pasport.jpg",
    "innUrl": "https://i.ibb.co/wF0ZrqwX/inn.jpg",
    "dmsUrl": "https://i.ibb.co/QxgKF0h/dms.jpg",
    "educationUrl": "https://i.ibb.co/6cHmLbdC/edu.jpg",
    "patentUrl": "https://i.ibb.co/vCQWwxs8/patent.jpg",
}

LIVE = store.DEFAULT_FOLDER / store.EXCEL_NAME


@pytest.fixture(autouse=True)
def isolated(monkeypatch):
    sandbox = tempfile.mkdtemp()
    monkeypatch.setenv("XDG_DATA_HOME", sandbox)
    monkeypatch.setenv("LOCALAPPDATA", sandbox)
    paths.data_dir.cache_clear()
    yield
    paths.data_dir.cache_clear()


@pytest.fixture()
def their_folder(tmp_path):
    """A COPY of the office's importer — their own is never written to here."""
    if not store.DEFAULT_FOLDER.exists():
        pytest.skip("АМИНА папкаси бу компютерда йўқ")
    copy = tmp_path / "amina"
    shutil.copytree(store.DEFAULT_FOLDER, copy,
                    ignore=shutil.ignore_patterns("node_modules", "~$*"))
    return copy


def _worker(**over) -> store.AminaData:
    made = store.AminaData(
        full_name="Шарипов Эхромиддин Талбишоевич", gender="Мужской",
        dob="20-01-1998", birth_place="Таджикистан",
        citizenship="Республика Таджикистан",
        phone="+7 996 681-84-92", extra_phone="+7 996 681-84-92",
        street="улица Беловежская", house="71", apartment="94",
        kig_number="АВ0461171", kig_expire="2031-06-05")
    for key, value in over.items():
        setattr(made, key, value)
    return made


def _photo(label: str = "DOC", tilt: float = 5,
           card: tuple[int, int] = (700, 460),
           desk: tuple[int, int] = (1200, 900)) -> bytes:
    """A card photographed on a desk: tilted, on a flat surface.

    The desk is deliberately a plain mid grey with no detail in it, and the
    card is covered in printing — because detail is what the cut goes by.
    """
    from PIL import Image, ImageDraw

    surface = Image.new("RGB", desk, (104, 99, 94))
    width, height = card
    paper = Image.new("RGB", card, (250, 249, 245))
    draw = ImageDraw.Draw(paper)
    draw.rectangle([0, 0, width - 1, height - 1], outline=(120, 130, 120),
                   width=4)
    draw.text((28, 26), label, fill=(20, 20, 20))
    for row in range(6):                       # printing, so it has texture
        y = int(height * 0.2) + row * int(height * 0.11)
        draw.line([(30, y), (width - 40, y)], fill=(60, 70, 85), width=5)
    paper = paper.rotate(tilt, expand=True, fillcolor=(104, 99, 94))
    surface.paste(paper, ((desk[0] - paper.width) // 2,
                          (desk[1] - paper.height) // 2))
    buf = io.BytesIO()
    surface.save(buf, "JPEG", quality=92)
    return buf.getvalue()


# ------------------------------------------------- the login and password
def test_the_password_is_the_phone_with_an_eight_in_front() -> None:
    """«+7 996 681-84-92» → «89966818492». The office's own rule."""
    assert store.password_of("+7 996 681-84-92") == "89966818492"
    assert store.password_of("+79966818492") == "89966818492"
    assert store.password_of("89966818492") == "89966818492"
    assert store.password_of("9966818492") == "89966818492"


def test_a_number_that_is_not_a_number_is_not_invented() -> None:
    """A made-up password opens nothing — better to hand back what there is."""
    assert store.password_of("") == ""
    assert store.password_of("абв") == ""
    assert store.password_of("+998 90 123-45-67") == "998901234567"


def test_the_login_is_the_surname_and_the_last_four_digits() -> None:
    assert store.email_of("Шарипов", "+7 996 681-84-92") == SAMPLE["email"]
    assert store.email_of("Эргешов", "+7 900 111-22-33") == "ergeshov2233@gmail.com"
    assert store.email_of("", "+7 996 681-84-92") == ""


def test_the_surname_is_written_in_latin_letters() -> None:
    assert store.latin("Шарипов") == "sharipov"
    assert store.latin("Хужаев") == "khujaev"
    assert store.latin("Ўринов") == "orinov"


def test_the_worker_carries_his_own_login_and_password() -> None:
    worker = _worker()
    assert worker.login() == SAMPLE["email"]
    assert worker.password() == SAMPLE["password"]


# ------------------------------------------------------------ the address
def test_the_address_line_is_built_the_way_the_office_writes_it() -> None:
    assert _worker().address_line() == SAMPLE["addressSettlement"]


def test_an_address_line_typed_by_hand_wins() -> None:
    typed = _worker(settlement="Москва, Зеленоград, корпус 1615")
    assert typed.address_line() == "Москва, Зеленоград, корпус 1615"


# -------------------------------------------------------- the citizenship
def test_the_citizenship_is_written_the_way_the_app_wants_it() -> None:
    assert store.citizenship_of("ТАДЖИКИСТАН") == "Республика Таджикистан"
    assert store.citizenship_of("Узбекистан") == "Республика Узбекистан"
    assert store.citizenship_of("Киргизия") == "Кыргызская Республика"
    assert store.citizenship_of("Россия") == "Российская Федерация"
    assert store.citizenship_of("") == ""


# ---------------------------------------------------------- the passport
@dataclass
class _Passport:
    surname: str = "Шарипов"
    name: str = "Эхромиддин"
    patronymic: str = "Талбишоевич"
    birth_date: date | None = date(1998, 1, 20)
    gender: str = "M"
    nationality: str = "ТАДЖИКИСТАН"
    birth_place: str = "Таджикистан"


def test_what_the_passport_gives_is_what_the_rows_want() -> None:
    read = store.data_of(_Passport())
    assert read.full_name == SAMPLE["fullName"]
    assert read.gender == SAMPLE["gender"]
    assert read.dob == SAMPLE["dob"], "кун-ой-йил бўлиши керак"
    assert read.birth_place == SAMPLE["birthPlace"]
    assert read.citizenship == SAMPLE["citizenship"]


def test_a_woman_is_read_as_one() -> None:
    assert store.data_of(_Passport(gender="F")).gender == "Женский"
    assert store.data_of(_Passport(gender="жен")).gender == "Женский"


# ------------------------------------------------------------- the rows
def test_the_rows_come_out_exactly_as_the_office_filled_them() -> None:
    """Their own sample, reproduced value for value."""
    worker = _worker(urls={k: SAMPLE[k] for k in store.DOCS})
    assert worker.rows() == SAMPLE


def test_a_row_with_no_picture_is_left_empty() -> None:
    """«нечта юкласам ўшалар бўлади» — the rest must not carry a stale link."""
    worker = _worker(urls={"identityUrl": SAMPLE["identityUrl"]})
    rows = worker.rows()
    assert rows["identityUrl"] == SAMPLE["identityUrl"]
    assert rows["innUrl"] == ""
    assert rows["dmsUrl"] == rows["educationUrl"] == rows["patentUrl"] == ""


def test_the_second_phone_falls_back_to_the_first() -> None:
    assert _worker(extra_phone="").rows()["extraPhone"] == SAMPLE["phone"]


# ------------------------------------------------------------- the sheet
def test_the_picture_is_big_enough_not_to_be_shrunk_by_the_app() -> None:
    """The app does not enlarge one smaller than its frame.

    The office's own uploads are 2480×3507; anything much under that sits in
    the middle of the frame looking shrunken, with the app's grey around it.
    """
    from PIL import Image

    with Image.open(io.BytesIO(store.sheet_jpeg([_photo()]))) as sheet:
        assert max(sheet.size) == store.LONG_SIDE == 3500
        assert sheet.mode == "RGB"


def test_the_picture_is_the_document_and_nothing_else() -> None:
    """«полный саҳифа катакни эгалайдиган қилибер, бўш жой қолмасин».

    No white page under it and no desk around it: every edge of the picture
    is document, so there is nothing for the app to letterbox.
    """
    import numpy as np
    from PIL import Image

    with Image.open(io.BytesIO(store.sheet_jpeg([_photo()]))) as sheet:
        grey = np.asarray(sheet.convert("L"))
    ink = grey < 235
    columns = np.where(ink.any(axis=0))[0]
    rows = np.where(ink.any(axis=1))[0]
    wide = (columns[-1] - columns[0]) / grey.shape[1]
    tall = (rows[-1] - rows[0]) / grey.shape[0]
    assert wide > 0.97, f"эни бўйича {wide:.0%} — ёнларида бўш жой бор"
    assert tall > 0.97, f"бўйи бўйича {tall:.0%} — тепа-пастида бўш жой бор"


def test_the_desk_a_document_was_photographed_on_is_cut_away() -> None:
    """Measured on the office's own passport: the wall scored 3.5 for local
    texture and the passport 19.0. The desk is what the cut is looking for."""
    import numpy as np
    from PIL import Image

    # a small card adrift in a wide grey desk — mostly background
    photo = _photo(card=(560, 380), desk=(1600, 1200))
    with Image.open(io.BytesIO(photo)) as before:
        was = before.size
    with Image.open(io.BytesIO(store.sheet_jpeg([photo]))) as after:
        became = after.size

    assert became[0] / became[1] > was[0] / was[1], "кенг стол кесилмади"
    grey = np.asarray(Image.open(io.BytesIO(store.sheet_jpeg([photo])))
                      .convert("L"))
    # the desk is a flat mid grey; almost none of it should be left
    desk = ((grey > 85) & (grey < 125)).mean()
    assert desk < 0.10, f"стол фонининг {desk:.0%} и қолиб кетди"


def test_a_document_that_fills_its_photograph_is_left_alone() -> None:
    """A crop that would take a passport down to its own portrait photo is
    worse than a margin of desk — so a small find is refused."""
    from PIL import Image

    flat = Image.new("RGB", (900, 1200), (247, 246, 244))
    buf = io.BytesIO()
    flat.save(buf, "JPEG", quality=92)
    made = store.sheet_jpeg([buf.getvalue()])
    with Image.open(io.BytesIO(made)) as sheet:
        assert sheet.width / sheet.height == pytest.approx(0.75, abs=0.06)


def test_an_id_cards_two_sides_share_one_picture() -> None:
    """«олди ва орқасини битта қоғозга» — one picture, both on it, and the
    two matched to the same width so neither looks the larger."""
    import numpy as np
    from PIL import Image

    made = store.sheet_jpeg([_photo("FRONT", 6), _photo("BACK", -4)])
    with Image.open(io.BytesIO(made)) as sheet:
        assert sheet.height > sheet.width, "иккови устма-уст турмаган"
        half = sheet.height // 2
        top = np.asarray(sheet.crop((0, 0, sheet.width, half)).convert("L"))
        bottom = np.asarray(
            sheet.crop((0, half, sheet.width, sheet.height)).convert("L"))
    assert top.min() < 200, "юқориги ярмида ҳужжат йўқ"
    assert bottom.min() < 200, "пастки ярмида ҳужжат йўқ"


def test_a_field_takes_at_most_two_pictures() -> None:
    with pytest.raises(ValidationError):
        store.sheet_jpeg([_photo(), _photo(), _photo()])
    with pytest.raises(ValidationError):
        store.sheet_jpeg([])


def test_only_the_identity_row_takes_two() -> None:
    assert store.doc_limit("identityUrl") == 2
    assert all(store.doc_limit(k) == 1 for k in store.DOCS
               if k != "identityUrl")


# ------------------------------------------------------------- the Excel
def test_the_excel_comes_out_matching_the_office_sample(their_folder) -> None:
    """Filled by us, read back — every row equal to what they typed."""
    excel = their_folder / store.EXCEL_NAME
    worker = _worker(urls={k: SAMPLE[k] for k in store.DOCS})
    store.fill_excel(worker.rows(), excel)
    assert store.read_excel(excel) == SAMPLE


def test_the_excel_keeps_its_own_shape(their_folder) -> None:
    """Both sheets, the row labels and the text formatting all survive.

    A date written as a DATE reaches the importer as a serial number and the
    worker's birthday lands in the app as «44911» — so every cell stays text.
    """
    from openpyxl import load_workbook

    excel = their_folder / store.EXCEL_NAME
    store.fill_excel(_worker().rows(), excel)

    book = load_workbook(excel)
    assert book.sheetnames == ["user", "how_to"]
    sheet = book["user"]
    assert sheet["A1"].value == "field" and sheet["B1"].value == "value"
    labels = [r[0].value for r in sheet.iter_rows(min_row=2) if r[0].value]
    assert labels == list(store.FIELDS)
    for row in sheet.iter_rows(min_row=2):
        if not row[0].value:
            continue
        # An empty row comes back as None — openpyxl stores no cell for "".
        # That is what a row with no picture SHOULD be: their importer reads
        # `if (!fileUrl) continue`, so the document is simply skipped.
        assert row[1].value is None or isinstance(row[1].value, str), \
            f"{row[0].value} матн эмас — {type(row[1].value).__name__}"
        assert row[1].number_format == "@"
    # and the dates in particular are text, not dates
    typed = {r[0].value: r[1].value for r in sheet.iter_rows(min_row=2)}
    assert typed["dob"] == "20-01-1998"
    assert typed["kigExpire"] == "2031-06-05"
    assert sum(1 for r in book["how_to"].iter_rows(values_only=True)
               if any(c is not None for c in r)) == 6


def test_writing_the_excel_touches_nothing_else_in_their_folder(
        their_folder) -> None:
    """«шу папкадаги ҳеч қайси файлга тегма» — held to, and proved."""
    def fingerprints() -> dict[str, str]:
        return {str(p.relative_to(their_folder)):
                hashlib.sha256(p.read_bytes()).hexdigest()
                for p in sorted(their_folder.rglob("*")) if p.is_file()}

    before = fingerprints()
    store.fill_excel(_worker().rows(), their_folder / store.EXCEL_NAME)
    after = fingerprints()

    assert set(before) == set(after), "файл қўшилди ёки ўчди"
    changed = [name for name in before if before[name] != after[name]]
    assert changed == [store.EXCEL_NAME], f"тегилмаслиги керак эди: {changed}"


def test_the_untouched_original_is_kept_somewhere_of_ours(
        their_folder) -> None:
    """Whatever the file said before, that is what the copy must say.

    NOT compared against the sample: the office is using this section now,
    so its live Excel holds whichever worker went through last. A test that
    expects a particular name in there breaks every time they run one.
    """
    excel = their_folder / store.EXCEL_NAME
    before = store.read_excel(excel)
    store.fill_excel(_worker().rows(), excel)

    kept = paths.user_templates_dir() / store.SECTION / f"original_{store.EXCEL_NAME}"
    assert kept.exists(), "асл нусха сақланмади"
    assert store.read_excel(kept) == before, "сақлангани асл эмас"
    assert store.read_excel(excel) != before, "эксел тўлмади"


def test_an_excel_missing_a_row_is_refused_not_half_written(
        their_folder) -> None:
    """Half a worker in Firebase is worse than none — so it stops first."""
    from openpyxl import load_workbook

    excel = their_folder / store.EXCEL_NAME
    book = load_workbook(excel)
    sheet = book["user"]
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == "kigNumber":
            row[0].value = "somethingElse"
    book.save(excel)

    with pytest.raises(ValidationError, match="kigNumber"):
        store.fill_excel(_worker().rows(), excel)


# ---------------------------------------------------------- the machinery
def test_the_office_is_told_what_is_missing_before_any_work(tmp_path) -> None:
    class _Settings:
        @staticmethod
        def get(key, default=None):
            return str(tmp_path) if key == store.KEY_FOLDER else default

    with pytest.raises(ValidationError, match="Эксел|папка"):
        store.check_folder(_Settings())


def test_a_stale_lock_file_does_not_block_the_office_for_ever(
        their_folder) -> None:
    """Excel leaves a ``~$`` file behind when it crashes — and the office has
    one, months older than the workbook. Going by it would refuse for ever."""
    class _Settings:
        @staticmethod
        def get(key, default=None):
            return str(their_folder) if key == store.KEY_FOLDER else default

    (their_folder / f"~${store.EXCEL_NAME}").write_bytes(b"stale")
    assert store.excel_is_open(_Settings()) is False, "эски қулфга қараб рад этди"


def test_an_excel_really_held_open_is_noticed(their_folder,
                                              monkeypatch) -> None:
    class _Settings:
        @staticmethod
        def get(key, default=None):
            return str(their_folder) if key == store.KEY_FOLDER else default

    assert store.excel_is_open(_Settings()) is False

    real = Path.open

    def denied(self, *args, **kw):
        if self.name == store.EXCEL_NAME and "r+b" in args:
            raise PermissionError(32, "held by Excel")
        return real(self, *args, **kw)

    monkeypatch.setattr(Path, "open", denied)
    assert store.excel_is_open(_Settings()) is True


def test_nothing_is_sent_anywhere_when_the_worker_is_incomplete(
        their_folder) -> None:
    """Refused first, so no picture is published for an account never made."""
    class _Settings:
        @staticmethod
        def get(key, default=None):
            return str(their_folder) if key == store.KEY_FOLDER else default

    excel = their_folder / store.EXCEL_NAME
    before = store.read_excel(excel)
    service = store.AminaService(_Settings())
    with pytest.raises(ValidationError, match="ФИО"):
        service.create(_worker(full_name=""), {}, run=False)
    with pytest.raises(ValidationError, match="Телефон"):
        service.create(_worker(phone=""), {}, run=False)
    # a refusal leaves the Excel exactly as it found it
    assert store.read_excel(excel) == before


def test_pictures_without_an_imgbb_key_are_refused_before_the_excel(
        their_folder) -> None:
    class _Settings:
        @staticmethod
        def get(key, default=None):
            return str(their_folder) if key == store.KEY_FOLDER else default

    service = store.AminaService(_Settings())
    with pytest.raises(ValidationError, match="imgbb"):
        service.create(_worker(), {"innUrl": [_photo()]}, run=False)


def test_a_worker_with_links_already_needs_no_key(their_folder) -> None:
    """Re-running a worker whose pictures are already up must still work."""
    class _Settings:
        @staticmethod
        def get(key, default=None):
            return str(their_folder) if key == store.KEY_FOLDER else default

    service = store.AminaService(_Settings())
    worker = _worker(urls={k: SAMPLE[k] for k in store.DOCS})
    result = service.create(worker, {}, run=False)
    assert result.login == SAMPLE["email"]
    assert result.password == SAMPLE["password"]
    assert store.read_excel(their_folder / store.EXCEL_NAME) == SAMPLE


def test_the_slip_is_the_two_lines_the_worker_needs() -> None:
    made = store.AminaResult(login="sharipov8492@gmail.com",
                             password="89966818492", urls={}, output="")
    assert made.slip() == ("Логин: sharipov8492@gmail.com\n"
                           "Парол: 89966818492")


# ---------------------------------------------------------- the importer
#: A stand-in for the office's own script. Their real one creates a Firebase
#: account, which is not a thing a test may do — but the plumbing around it
#: (both commands, in order, output captured, failure surfaced) is, and that
#: is what these check.
def _fake_importer(where: Path, says: str, code: int = 0) -> None:
    where.mkdir(parents=True, exist_ok=True)
    (where / "package.json").write_text(
        '{"name":"fake","version":"1.0.0","scripts":'
        f'{{"import":"node -e \\"console.log(\'{says}\');'
        f'process.exit({code})\\""}}}}', "utf-8")


def test_both_commands_run_in_order_and_their_output_comes_back(
        tmp_path) -> None:
    """«иккта буйруқ бор — инстал ва импорт, иккаласини ҳам»."""
    where = tmp_path / "importer"
    _fake_importer(where, "IMPORT FINISHED")
    said = store.run_import(where, timeout=120)
    assert "$ npm install" in said
    assert "$ npm run import" in said
    assert said.index("npm install") < said.index("npm run import")
    assert "IMPORT FINISHED" in said


def test_a_failing_import_is_surfaced_not_swallowed(tmp_path) -> None:
    """An e-mail already registered must reach the office, not vanish."""
    where = tmp_path / "importer"
    _fake_importer(where, "auth/email-already-exists", code=1)
    with pytest.raises(OfisError, match="npm run import"):
        store.run_import(where, timeout=120)


def test_an_import_that_never_says_it_finished_is_not_called_a_success(
        tmp_path) -> None:
    where = tmp_path / "importer"
    _fake_importer(where, "started, then nothing")
    with pytest.raises(OfisError, match="тугамади"):
        store.run_import(where, timeout=120)


# ------------------------------------------------- what the office types
def test_the_city_and_second_phone_are_still_there_next_time() -> None:
    assert store.typed() == {}
    store.remember_typed(city="Москва", extra_phone="+7 996 681-84-92")
    assert store.typed() == {"city": "Москва",
                             "extra_phone": "+7 996 681-84-92"}
    store.remember_typed(city="Химки")
    assert store.typed()["city"] == "Химки"
    assert store.typed()["extra_phone"] == "+7 996 681-84-92"


# --------------------------------------------- the office's own live file
def test_the_live_folder_is_never_written_to_by_the_test_suite() -> None:
    """A guard on ourselves: their real Excel must still say what it said."""
    if not LIVE.exists():
        pytest.skip("АМИНА папкаси бу компютерда йўқ")
    assert store.read_excel(LIVE)["fullName"], "жонли эксел бузилган"
